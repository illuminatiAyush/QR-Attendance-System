from flask import Flask, request, send_file, render_template
import openpyxl
import os
import qrcode
from datetime import datetime

app = Flask(__name__)

# Function to get the Excel file name with the current date
def get_excel_filename():
    today = datetime.now().strftime("%Y-%m-%d")  # Format: YYYY-MM-DD
    return f"data_{today}.xlsx"

# Ensure the Excel file exists
def ensure_excel_file():
    excel_file = get_excel_filename()
    if not os.path.exists(excel_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["IP Address", "Name", "Email"])
        wb.save(excel_file)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/capture", methods=["POST"])
def capture():
    ip_address = request.remote_addr
    name = request.form.get("name")
    email = request.form.get("email")

    excel_file = get_excel_filename()
    ensure_excel_file()

    # Open the Excel file
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # Check if IP already exists
    ip_exists = any(ws.cell(row=i, column=1).value == ip_address for i in range(2, ws.max_row + 1))

    if not ip_exists:
        ws.append([ip_address, name, email])
        wb.save(excel_file)
        return "Data saved successfully!", 200
    else:
       return "IP address already exists!", 400

@app.route("/download")
def download():
    excel_file = get_excel_filename()
    
    return send_file(excel_file, as_attachment=True)

if __name__ == "__main__":
    # Generate QR code only once when the server starts
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data("http://192.168.1.16:5000")
    qr.make(fit=True)

    # Create an image from the QR code
    img = qr.make_image(fill="black", back_color="white")
    img.save("qrcode.png")

    app.run(host="192.168.1.16", port=5000)

